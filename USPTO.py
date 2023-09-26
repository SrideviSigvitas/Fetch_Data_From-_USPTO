from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import time
import requests
import select
import PyPDF2
import os
import io
import openpyxl
import shutil
from selenium.webdriver.support.ui import Select
workbook=openpyxl.load_workbook(r"C:\Users\DELL\Desktop\Law Firm Index - 15 Sep 2023 - Sridevi - Kilpatrick 610 to 690.xlsx")
activeSheet=workbook.active
rows=activeSheet.max_row
columns=activeSheet.max_column

cell_value1=activeSheet.cell(row=4,column=2).value
print("Patent Number-->-->",cell_value1)
cell_value2=activeSheet.cell(row=4,column=3).value
print("Priority-->-->",cell_value2)
cell_value3=activeSheet.cell(row=4,column=4).value
print("Filed-->-->",cell_value3)
cell_value4=activeSheet.cell(row=4,column=5).value
print("Date of Paten-->",cell_value4)
cell_value5=activeSheet.cell(row=4,column=6).value
print("Reverse Citation-->",cell_value5)
cell_value6=activeSheet.cell(row=4,column=7).value
print("Signing Attorney-->",cell_value6)
cell_value7=activeSheet.cell(row=4,column=8).value
print("Patent agent Number-->",cell_value7)
cell_value8=activeSheet.cell(row=4,column=9).value
print("CPC-->",cell_value8)
cell_value9=activeSheet.cell(row=4,column=10).value
print("CPC Definition-->",cell_value9)
cell_value10=activeSheet.cell(row=4,column=11).value
print("Title-->",cell_value10)
cell_value11=activeSheet.cell(row=4,column=12).value
print("Assignee-->",cell_value10)


options=webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver=webdriver.Chrome(options=options)
driver.maximize_window()
driver.get("https://patentcenter.uspto.gov/")
driver.implicitly_wait(5)
Search_Box = driver.find_element(By.XPATH,"//input[@id='TxtBox_bibData_search_input']")
print("Patent Number-->",cell_value1)
Search_Box.send_keys(cell_value1)
drop_down=driver.find_element(By.XPATH,"//button[@class='btn btn-primary dropdown-toggle']")
drop_down.click()
driver.implicitly_wait(5)
Select_Patents=driver.find_element(By.XPATH,"(//a[@class='dropdown-item'])[18]")
Select_Patents.click()
#print(p)
Search_Box=driver.find_element(By.XPATH,"//button[@type='submit']")
Search_Box.click()

Docment_Tansactions=driver.find_element(By.XPATH,"//a[@id='link_app-data-docs-trans']")
Docment_Tansactions.click()
Ads=driver.find_element(By.XPATH,"//*[@id='DataTables_Table_0']/tbody/tr[191]/td[3]")
driver.execute_script("arguments[0].scrollIntoView();",Ads)
Click_PDF=driver.find_element(By.XPATH,"//*[@id='DataTables_Table_0']/tbody/tr[191]/td[6]/button")
Click_PDF.click()
print("PDF File dowloaded")
driver.implicitly_wait(5)

driver.get("file:///C:/Users/DELL/Downloads/20350_15888979_02-06-2018_ADS%20(2).PDF")
Download=driver.find_element(By.XPATH,"//*[@id='print']")
Download.click()

# requests.get(url)
# file_name = '11672694.PDF'
# driver.implicitly_wait(5)
# #requests.get("file:///C:/Users/DELL/Downloads/20350_15888979_02-06-2018_ADS%20(2).PDF")
# response = requests.get(url)
# with open(r'C:\Users\DELL\Desktop\Python_Automated_PDF\11672694.PDF', 'wb') as file:
#     driver.implicitly_wait(5)
# file.write(response.content)
 

  
#PDF=driver.find_element(By.XPATH,"//*[@id='file-link']")
#PDF.click()